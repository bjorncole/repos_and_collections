# A way of putting classes into Excel and vice versa

import xlwings as xl
import inspect
import networkx as nx
from repos_and_collections.collections import Collection
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.comments import Comment

class ExcelCollectionSet(object):
    '''
    Represents Excel workbooks as a set of collections, one per sheet
    '''

    def __init__(self):
        self.collection_list = []

    def write_to_excel(self):
        '''
        Writes to Excel as a graph, noting the path should ultimately terminate in individual values or a list
        (For example, tree of merged cells on the left via rows and lists expanding column-wise or the transpose)
        :return: None (side effect is a modified Excel sheet)
        '''

        max_graph_depth = 0

class ExcelCollection(Collection):

    '''
    Represents a collection of Excel objects (serializes as a single sheet)
    '''

    def __init__(self):
        super().__init__()
        # if true, put graph on first columns of sheet, if false put graph
        # on first rows of sheet
        self.row_headers = False
        self.graphs = []
        self.roots = []
        self.cell_dict = {}
        self.aux_graph_dict = {}
        self.mapped_sheet = None

    def add_graph(self, incoming_graph, incoming_root, incoming_aux):
        new_graph = ExcelCellPath()
        new_graph.path_graph = incoming_graph
        self.graphs.append(new_graph)
        self.roots.append(incoming_root)
        self.aux_graph_dict.update(incoming_aux)

    def style_range(self, ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

    def build_cell_dict(self):
        total_max_depth = 0
        total_width = 0
        col_counter = 1

        for indx, grph in enumerate(self.graphs):
            # get max depth and width for a given graph

            leaf_list = []

            for key, val in nx.out_degree_centrality(grph.path_graph).items():
                if val == 0.0:
                    leaf_list.append(key)

            # assume row-wise to start; can work transpose later

            for leaf in leaf_list:
                for tree_path in list(nx.all_simple_paths(grph.path_graph, self.roots[0], leaf)):
                    path_builder = ''
                    for indx, step in enumerate(tree_path):
                        if indx == 0:
                            path_builder = step
                        else:
                            path_builder = path_builder + '.' + step

                    grph.path_dict.update({path_builder: tree_path})

            for path in sorted(grph.path_dict):
                if total_max_depth < len(grph.path_dict[path]):
                    total_max_depth = len(grph.path_dict[path])
                row_counter = len(grph.path_dict[path])
                for indx2, step in enumerate(reversed(grph.path_dict[path])):

                    path_length = len(grph.path_dict[path])

                    print(grph.path_dict[path][path_length - indx2 - 2])
                    print(step)

                    test_dat = grph.path_graph.\
                        get_edge_data(grph.path_dict[path][path_length - indx2 - 2],
                                      step)

                    print('Edge data = ' + str(test_dat))

                    if test_dat is not None:
                        self.aux_graph_dict.update({step: test_dat})

                    if step in self.cell_dict:
                        current_list = self.cell_dict[step]
                        if not self.row_headers:
                            current_list.append([row_counter, col_counter])
                        else:
                            current_list.append([col_counter, row_counter])
                    else:
                        if not self.row_headers:
                            self.cell_dict.update({step: [[row_counter, col_counter]]})
                        else:
                            self.cell_dict.update({step: [[col_counter, row_counter]]})

                    row_counter = row_counter - 1

                col_counter = col_counter + 1

            total_width = col_counter - 1

            print('Max depth on cells = ' + repr(total_max_depth))
            print('Width of cells = ' + repr(total_width))

            print(repr(self.cell_dict))

            grph.leaf_list = leaf_list

    def build_sheet(self, sht, aux_uml_graph):
        # Populate the Excel sheet
        # TODO: Fix leaky abstraction due to working with UML

        dark_line = Side(border_style="thin", color="000000")

        cell_description = ""

        print(repr(self.aux_graph_dict))

        for label in self.cell_dict:
            # if cells are next to each other, merge
            merge_start = 1
            merge_end = 0
            last_incr = 0
            current_level = 0

            # holders for values to support transposing
            row_a = 0
            row_b = 0
            col_a = 0
            col_b = 0

            if label == '_id':
                cell_description = 'Place ID here to reference another row'
            else:
                cell_description = ''
            for cell_assignment in self.cell_dict[label]:
                incr = 0
                level = 0

                if self.row_headers:
                    level = cell_assignment[1]
                    incr = cell_assignment[0]
                else:
                    level = cell_assignment[0]
                    incr = cell_assignment[1]

                current_level = level
                # if cell locations are adjacent, track them for merging
                if incr - last_incr == 1 and incr != 1:
                    merge_end = incr
                # once adjacency stops, go ahead and merge
                else:
                    if merge_end > 0 and not label in self.graphs[0].leaf_list:
                        if self.row_headers:
                            row_a = merge_start
                            row_b = merge_end
                            col_a = current_level
                            col_b = current_level
                        else:
                            row_a = current_level
                            row_b = current_level
                            col_a = merge_start
                            col_b = merge_end

                        print(
                            'Merging range ' +
                            "{0}".format(get_column_letter(col_a)) + repr(row_a) + ":" +
                            "{0}".format(get_column_letter(col_b)) + repr(row_b) + ' for label ' + label)
                        sht.merge_cells(start_row=row_a, end_row=row_b, start_column=col_a,
                                        end_column=col_b)

                        self.style_range(sht,
                                    "{0}".format(get_column_letter(col_a)) + repr(row_a) + ":" +
                                    "{0}".format(get_column_letter(col_b)) + repr(row_b),
                                    border=Border(left=dark_line, right=dark_line, top=dark_line, bottom=dark_line)
                                    )
                    merge_end = 0
                    merge_start = incr
                sht["{0}".format(get_column_letter(cell_assignment[1])) + repr(cell_assignment[0])] = label

                comment_field = {}
                comment_field.update(self.aux_graph_dict[label])
                comment_field.update({'description': cell_description})

                sht["{0}".format(get_column_letter(cell_assignment[1])) +
                    repr(cell_assignment[0])].comment = \
                    Comment(repr(comment_field).replace('\'', '\"'), 'autobuilder')
                    #Comment('{\"type\": \"' + aux_uml_graph.type_dict[label].name +
                    #        '\",\n \"description\": \"' + cell_description + '\"}', 'autobuilder')

                if self.row_headers:
                    sht["{0}".format(get_column_letter(cell_assignment[1]))
                        + repr(cell_assignment[0])].alignment = Alignment(vertical='center')
                else:
                    sht["{0}".format(get_column_letter(cell_assignment[1]))
                        + repr(cell_assignment[0])].alignment = Alignment(horizontal='center')
                sht["{0}".format(get_column_letter(cell_assignment[1]))
                    + repr(cell_assignment[0])].font = Font(bold=True)

                if label in self.graphs[0].leaf_list:
                    sht["{0}".format(get_column_letter(cell_assignment[1]))
                        + repr(cell_assignment[0])].border = \
                        Border(left=dark_line, right=dark_line, top=dark_line, bottom=dark_line)

                last_incr = incr
            # section to catch case where merge is at the end of a line
            if merge_end > 0 and not label in self.graphs[0].leaf_list:

                if self.row_headers:
                    row_a = merge_start
                    row_b = merge_end
                    col_a = current_level
                    col_b = current_level
                else:
                    row_a = current_level
                    row_b = current_level
                    col_a = merge_start
                    col_b = merge_end

                print(
                    'Merging range ' +
                    "{0}".format(get_column_letter(col_a)) + repr(row_a) + ":" +
                    "{0}".format(get_column_letter(col_b)) + repr(row_b) + ' for label ' + label)

                sht.merge_cells(start_row=row_a, end_row=row_b, start_column=col_a,
                                end_column=col_b)

                self.style_range(sht,
                            "{0}".format(get_column_letter(col_a)) + repr(row_a) + ":" +
                            "{0}".format(get_column_letter(col_b)) + repr(row_b),
                            border=Border(left=dark_line, right=dark_line, top=dark_line, bottom=dark_line))
            else:
                sht["{0}".format(get_column_letter(cell_assignment[1]))
                    + repr(current_level)].border = \
                    Border(left=dark_line, right=dark_line, top=dark_line, bottom=dark_line)

    def read_header_to_graph(self):
        '''
        Look at a given Excel sheet and convert it into a structural graph

        :return: None (updates internal graph)
        '''

class ExcelCellPath(object):
    '''
    A convenience class for calculating and holding the paths for value locations
    '''

    def __init__(self):
        self.path_graph = None
        self.type_dict = {}
        self.leaf_list = []
        self.path_dict = {}