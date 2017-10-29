from repos_and_collections.excel_objects import ExcelCollectionSet
from pprint import PrettyPrinter
from deep_metamodeling.metamodels.m2_uml import Class, Property, ValueType, CompositionGraph
from repos_and_collections.collections import Collection
from repos_and_collections.excel_objects import ExcelCollection
import networkx as nx
import openpyxl as xl
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.comments import Comment
from openpyxl import load_workbook
import inspect
import simplejson as json

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def main():

    '''
    A test of bringing a structured model in from Excel based on a pre-defined tree structure
    :return:
    '''

    # So - how to do this? Should each class get its own section? How to know when to nest and when
    # to do more of a cross-reference? Also need to be careful to avoid recursion since UML graphs
    # often have many paths to get to a given attribute

    xl_collect = ExcelCollectionSet()

    pp = PrettyPrinter(indent=2)

    xl_collection = ExcelCollection()
    xl_collection.name = 'Car Data'

    wb = load_workbook('TestWorkbook_top.xlsx')

    sht = wb.get_sheet_by_name(xl_collection.name)

    row_headers = False
    header_width = 0
    max_depth = 0
    max_row = 0
    max_col = 0

    base_label = ''

    type_dict = {}

    working_graph = nx.DiGraph()

    location_dict = {}

    for rng in sht.merged_cell_ranges:
        print(repr(coordinate_to_tuple(rng.split(':')[0])) + ' through ' + repr(coordinate_to_tuple(rng.split(':')[1])))
        merge_start = coordinate_to_tuple(rng.split(':')[0])
        merge_end = coordinate_to_tuple(rng.split(':')[1])

        label = sht[rng.split(':')[0]].value

        if merge_end[0] != merge_start[0]:
            for stp in range(merge_start[0], merge_end[0] + 1):
                location_dict.update({get_column_letter(merge_start[1]) + repr(stp): label})

        elif merge_end[1] != merge_start[1]:
            for stp in range(merge_start[1], merge_end[1] + 1):
                location_dict.update({get_column_letter(stp) + repr(merge_start[0]) : label})

        if max_row < merge_end[0]:
            max_row = merge_end[0]
        if max_col < merge_end[1]:
            max_col = merge_end[0]

        if merge_start[0] == 1 and merge_start[1] == 1:
            # need to detect if the headers are top columns or left-most rows
            if merge_end[0] != merge_start[0]:
                print('Header is across rows.')
                header_width = merge_end[0] - merge_start[0]
                row_headers = True
            elif merge_end[1] != merge_start[1]:
                print('Header is across columns.')
                header_width = merge_end[1] - merge_start[1]
                row_headers = False

        trial = sht[get_column_letter(merge_start[1]) + repr(merge_start[0])].comment.text

        comment_dict = json.loads(trial)

        type_dict.update({label: comment_dict['type_']})

    if row_headers:
        max_depth = max_col
    else:
        max_depth = max_row

    print('Width of header is ' + repr(header_width))
    print('Max depth from merge is ' + repr(max_depth))

    # generate a graph of the cells, pass on to Excel Collection for further processing

    col = 0
    row = 0

    for stp in range(1, header_width + 2):
        prev_label = ''
        found_something = False
        # guess for a decent number of levels to look for unmerged cells
        for lvl in range(1, max_depth + 10):
            if row_headers:
                row = stp
                col = lvl
            else:
                col = stp
                row = lvl
            # walk down the merged cells, then look for those with comments
            current_label = None

            if get_column_letter(col) + repr(row) in location_dict:
                current_label = location_dict[get_column_letter(col) + repr(row)]

            if current_label is not None:
                if row == 1 and col == 1:
                    base_label = current_label
                if prev_label != '':
                    working_graph.add_edge(prev_label,
                                           current_label, type_=type_dict[current_label])
                found_something = True
                prev_label = current_label
            else:
                trial = sht[get_column_letter(col) + repr(row)].comment
                #print('Comment at ' + repr(get_column_letter(col) + repr(row)) + ' is ' + repr(trial))

                if trial is None and found_something:
                    break
                elif trial is not None and found_something:
                    current_label = sht[get_column_letter(col) + repr(row)].value
                    location_dict.update({
                        get_column_letter(col) + repr(row):
                            sht[get_column_letter(col) + repr(row)].value})

                    trial_text = trial.text

                    comment_dict = json.loads(trial_text)

                    if 'type_' in comment_dict:
                        working_graph.add_edge(prev_label,
                                               current_label, type_=comment_dict['type_'])
                        type_dict.update({current_label: comment_dict['type_']})

                    elif 'valtype_' in comment_dict:
                        working_graph.add_edge(prev_label,
                                               current_label, valtype_=comment_dict['valtype_'])
                        type_dict.update({current_label: comment_dict['valtype_']})

                    prev_label = current_label

    print('Location dict is:')
    print(repr(location_dict))

    print('Type dict is:')
    print(repr(type_dict))

    print(repr(working_graph.edges()))

    class_collection = {}
    property_collection = []

    root_class = Class()
    root_class.name = base_label

    class_collection.update({root_class.name: root_class})

    # first pass for classes, then next pass for properties

    for edg in working_graph.edges():
        edge_data = working_graph.get_edge_data(edg[0],edg[1])
        class_to_apply = None

        for cls in class_collection:
            if 'type_' in edge_data:
                if cls == edge_data['type_']:
                    # apply class type to new property
                    class_to_apply = cls
            elif 'valtype_' in edge_data:
                if cls == edge_data['valtype_']:
                    # apply class type to new property
                    class_to_apply = cls
        if class_to_apply is None:
            if 'type_' in edge_data:
                new_class = Class()
                new_class.name = edge_data['type_']
                class_collection.update({new_class.name: new_class})
            if 'valtype_' in edge_data:
                new_class = ValueType()
                new_class.name = edge_data['valtype_']
                class_collection.update({new_class.name: new_class})

        print(working_graph.get_edge_data(edg[0],edg[1]))

    for edg in working_graph.edges():
        # check to see if parent class already has such a property

        parent_type = class_collection[type_dict[edg[0]]]
        child_type = class_collection[type_dict[edg[1]]]

        pre_exist = False

        for attr in parent_type.owned_attribute:
            if attr.name == edg[1]:
                pre_exist = True

        if not pre_exist:
            new_prop = Property()
            new_prop.name = edg[1]
            parent_type.owned_attribute.append(new_prop)
            new_prop.type_.append(child_type)
            property_collection.append(new_prop)


    print(repr(class_collection))

    for prp in property_collection:
        print('Prop collection contains prop ' + prp.name + ' of type '
              + prp.type_[0].name + ' owned by ' + prp.class_.name)

if __name__ == "__main__":
    # run test from command line

    main()

